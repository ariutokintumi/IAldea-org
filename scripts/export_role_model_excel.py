#!/usr/bin/env python3
"""
Exporta cada pestaña del libro Excel a CSV UTF-8 en docs/roles/.

Requisito: pip install openpyxl (por ejemplo en un venv del proyecto).

Uso desde la raíz del repo:
  python3 scripts/export_role_model_excel.py
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError as e:  # pragma: no cover
    print("Instala openpyxl: python3 -m venv .venv && .venv/bin/pip install openpyxl", file=sys.stderr)
    raise SystemExit(1) from e

REPO = Path(__file__).resolve().parents[1]
XLSX = REPO / "docs" / "Role Model + Permission Matrix - IAldea.xlsx"
OUT_DIR = REPO / "docs" / "roles"


def is_empty(c) -> bool:
    if c is None:
        return True
    if isinstance(c, str) and not c.strip():
        return True
    return False


def cell_str(c) -> str:
    if c is None:
        return ""
    if isinstance(c, float) and c == int(c):
        return str(int(c))
    return str(c).strip() if isinstance(c, str) else str(c)


def sheet_matrix(ws) -> list[list[object]]:
    return [list(row) for row in ws.iter_rows(values_only=True)]


def tight_bbox(rows: list[list[object]]) -> tuple[int, int, int, int] | None:
    """(r0, r1, c0, c1) inclusive indices of non-empty cells."""
    best = None
    for i, row in enumerate(rows):
        for j, c in enumerate(row):
            if not is_empty(c):
                if best is None:
                    best = [i, i, j, j]
                else:
                    best[0] = min(best[0], i)
                    best[1] = max(best[1], i)
                    best[2] = min(best[2], j)
                    best[3] = max(best[3], j)
    if best is None:
        return None
    return best[0], best[1], best[2], best[3]


def crop(rows: list[list[object]]) -> list[list[object]]:
    bb = tight_bbox(rows)
    if bb is None:
        return []
    r0, r1, c0, c1 = bb
    out = []
    for i in range(r0, r1 + 1):
        row = rows[i] if i < len(rows) else []
        seg = []
        for j in range(c0, c1 + 1):
            seg.append(row[j] if j < len(row) else None)
        out.append(seg)
    return out


def write_csv(path: Path, matrix: list[list[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for row in matrix:
            w.writerow([cell_str(c) for c in row])


def fill_rol_column_user_stories(matrix: list[list[object]]) -> list[list[object]]:
    """Cabecera Rol,ID,... y rellenar Rol en filas de continuación."""
    if not matrix:
        return matrix
    m = [list(r) for r in matrix]
    hdr = m[0]
    if len(hdr) > 1 and is_empty(hdr[0]) and cell_str(hdr[1]) == "ID":
        hdr[0] = "Rol"
    try:
        rol_idx = [cell_str(x) for x in hdr].index("Rol")
    except ValueError:
        rol_idx = 0
    last_rol = ""
    for i in range(1, len(m)):
        row = m[i]
        while len(row) < len(hdr):
            row.append(None)
        rcell = row[rol_idx] if rol_idx < len(row) else None
        if not is_empty(rcell):
            last_rol = cell_str(rcell)
        elif last_rol:
            row[rol_idx] = last_rol
    return m


EXCEL_NAME = "Role Model + Permission Matrix - IAldea.xlsx"


def read_csv_rows(path: Path) -> list[list[str]]:
    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.reader(f))


def esc_md(s: str) -> str:
    return (s or "").replace("|", "\\|").replace("\n", "<br>")


def write_roles_por_etapa(src: Path, dest: Path) -> None:
    rows = read_csv_rows(src)
    if len(rows) < 4:
        return
    headers = rows[1]
    col_desc = rows[2]
    role_cols = list(range(10, 17))

    lines: list[str] = []
    lines.append("# Roles por etapa (derivado de la matriz)")
    lines.append("")
    lines.append("Este documento **no sustituye** el CSV; resume la columna narrativa por rol en cada etapa del ciclo de coordinación.")
    lines.append("")
    lines.append("- **Export completo (hoja *Role Model + Permission Matrix*):** [role-model-permission-matrix-full.csv](role-model-permission-matrix-full.csv)")
    lines.append("- **Matriz canónica reducida (V1, sin columnas UX):** [permission-matrix.csv](permission-matrix.csv)")
    lines.append("")
    lines.append("## Glosario de columnas (fila guía del export)")
    lines.append("")
    lines.append("| Columna | Descripción (export) |")
    lines.append("|---------|----------------------|")
    lines.append("| **Etapa** | Nombre de la etapa del ciclo de coordinación |")
    for i, h in enumerate(headers):
        h = (h or "").strip()
        d = (col_desc[i] if i < len(col_desc) else "") or ""
        d = d.strip()
        if i == 0 and not d:
            continue
        if not h and not d:
            continue
        if h == "Etapa" and d:
            continue
        if h == "Preguntas clave para uX uI":
            d = "Preguntas clave para UX/UI"
        elif h == "Respuestas":
            d = "Respuesta por pregunta (completar en la hoja cuando aplique; a menudo vacía en export)."
        if d:
            lines.append(f"| **{h or '(vacío)'}** | {d} |")
    lines.append("")
    lines.append("## Ciclo: responsabilidades por rol")
    lines.append("")

    for r in rows[3:11]:
        if not r or not (r[0] or "").strip():
            continue
        etapa = r[0].strip()
        desc = (r[1] or "").strip()
        coord = (r[2] or "").strip()
        lines.append(f"### {etapa}")
        lines.append("")
        if desc:
            lines.append(f"*{desc}*")
            lines.append("")
        if coord:
            lines.append(f"- **Objeto de coordinación:** {coord}")
            lines.append("")
        lines.append("| Rol | Qué hace en esta etapa (matriz) |")
        lines.append("|-----|--------------------------------|")
        for j in role_cols:
            rn = (headers[j] if j < len(headers) else "").strip()
            cell = (r[j] if j < len(r) else "") or ""
            cell = cell.strip()
            lines.append(f"| {rn} | {esc_md(cell)} |")
        lines.append("")

    lines.append("## Bloque UX pendiente en el export (plantilla)")
    lines.append("")
    lines.append("Las últimas filas del CSV reservan dimensiones de producto (**Qué ve**, **Qué puede hacer**, etc.) por rellenar en diseño.")
    lines.append("")
    dest.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_user_stories_matriz(src: Path, dest: Path) -> None:
    rows = read_csv_rows(src)
    if len(rows) < 4:
        return
    us: list[str] = []
    us.append("# User stories y criterios — matriz de etapas")
    us.append("")
    us.append("Derivado de [role-model-permission-matrix-full.csv](role-model-permission-matrix-full.csv) (columnas *Preguntas clave para uX uI* y *Respuestas*).")
    us.append("")
    us.append("Historias por rol: [user-stories-por-rol.md](user-stories-por-rol.md). Historias cortas del taller: [user-stories.md](user-stories.md).")
    us.append("")

    for r in rows[3:11]:
        if not r or not (r[0] or "").strip():
            continue
        etapa = r[0].strip()
        desc = (r[1] or "").strip()
        preg = (r[17] if len(r) > 17 else "") or ""
        preg = preg.strip()
        resp = (r[18] if len(r) > 18 else "") or ""
        resp = resp.strip()
        us.append(f"## Etapa: **{etapa}**")
        us.append("")
        if desc:
            us.append(f"**Contexto:** {desc}")
            us.append("")
        us.append("**Como** comunidad / comité / ciudadanía **quiero** que la interfaz apoye esta etapa del ciclo **para** alinear expectativas con la gobernanza humana descrita en la matriz.")
        us.append("")
        if preg:
            us.append("**Criterios / preguntas de diseño (export):**")
            us.append("")
            for part in preg.replace("\r\n", "\n").split("\n"):
                part = part.strip()
                if part:
                    us.append(f"- {part}")
            us.append("")
        else:
            us.append("*(Sin preguntas listadas en el export para esta etapa.)*")
            us.append("")
        if resp:
            us.append("**Respuestas en export:**")
            us.append("")
            us.append(resp)
            us.append("")
        else:
            us.append("**Respuestas en export:** *(pendiente en la hoja — columna Respuestas vacía.)*")
            us.append("")

    dest.write_text("\n".join(us) + "\n", encoding="utf-8")


def write_user_stories_por_rol_md(csv_path: Path, dest: Path) -> None:
    rows = read_csv_rows(csv_path)
    if not rows:
        return
    hdr = rows[0]
    lines: list[str] = []
    lines.append("# User stories por rol")
    lines.append("")
    xlsx_url = EXCEL_NAME.replace(" ", "%20")
    lines.append(f"Generado desde la pestaña **UserStories por rol** del libro [{EXCEL_NAME}](../{xlsx_url}) (export CSV en repo).")
    lines.append("")
    lines.append("La columna **Rol** está repetida en cada fila para facilitar filtros e importaciones.")
    lines.append("")
    lines.append("- **CSV:** [user-stories-por-rol.csv](user-stories-por-rol.csv)")
    lines.append("- **Historias cortas del taller Día 2:** [user-stories.md](user-stories.md)")
    lines.append("- **Preguntas UX por etapa:** [user-stories-matriz.md](user-stories-matriz.md)")
    lines.append("")
    lines.append("Los roles **Admin técnico** y **Operador de Piloto** amplían el núcleo en [role-model.md](role-model.md).")
    lines.append("")

    idx = {name: i for i, name in enumerate(hdr)}
    cur_rol = ""
    for r in rows[1:]:
        while len(r) < len(hdr):
            r.append("")
        rol = r[idx.get("Rol", 0)] or ""
        sid = r[idx.get("ID", 1)] or ""
        story = r[idx.get("User story", 2)] or ""
        obj_ = r[idx.get("Objetivo", 3)] or ""
        priv = r[idx.get("Privacidad", 4)] or ""
        fuente = r[idx.get("Fuente requerida", 5)] or ""
        riesgo = r[idx.get("Riesgo UX / Safety", 6)] or ""
        req = r[idx.get("Requisito UX", 7)] or ""
        if rol.strip():
            cur_rol = rol.strip()
            lines.append(f"## {cur_rol}")
            lines.append("")
        lines.append(f"### {sid}")
        lines.append("")
        lines.append(f"**User story:** {story}")
        lines.append("")
        lines.append("| Campo | Contenido |")
        lines.append("|-------|-----------|")
        lines.append(f"| Objetivo | {esc_md(obj_)} |")
        lines.append(f"| Privacidad | {esc_md(priv)} |")
        lines.append(f"| Fuente requerida | {esc_md(fuente)} |")
        lines.append(f"| Riesgo UX / Safety | {esc_md(riesgo)} |")
        lines.append(f"| Requisito UX | {esc_md(req)} |")
        lines.append("")

    dest.write_text("\n".join(lines) + "\n", encoding="utf-8")


def csv_to_md_table(title: str, csv_path: Path, dest: Path, intro: str) -> None:
    rows = read_csv_rows(csv_path)
    if not rows:
        return
    lines = [f"# {title}", "", intro, ""]
    hdr = rows[0]
    lines.append("| " + " | ".join(esc_md(c) for c in hdr) + " |")
    lines.append("|" + "|".join(["---"] * len(hdr)) + "|")
    for r in rows[1:]:
        r = r + [""] * (len(hdr) - len(r))
        lines.append("| " + " | ".join(esc_md(r[i]) for i in range(len(hdr))) + " |")
    lines.append("")
    dest.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_excel_sheets_index(dest: Path) -> None:
    xlsx_url = EXCEL_NAME.replace(" ", "%20")
    xlsx_link = f"[{EXCEL_NAME}](../{xlsx_url})"
    body = f"""# Índice — hojas del libro Excel

Libro: {xlsx_link} (todas las pestañas). Los CSV se regeneran con `scripts/export_role_model_excel.py` (requiere `openpyxl`).

| Pestaña en Excel | Archivo CSV | Notas breves |
|------------------|-------------|----------------|
| Role Model + Permission Matrix | [role-model-permission-matrix-full.csv](role-model-permission-matrix-full.csv) | Matriz etapa × rol + UX + plantilla. |
| V1 Role Model + Permission Matr | [v1-role-model-permission-matrix.csv](v1-role-model-permission-matrix.csv) | Matriz sin columnas UX; copia a [permission-matrix.csv](permission-matrix.csv). |
| Role experience Matrix | [role-experience-matrix.csv](role-experience-matrix.csv) | Necesidad por rol, riesgos, componentes UX. |
| Matriz de comportamiento por ro | [matriz-comportamiento-por-rol.csv](matriz-comportamiento-por-rol.csv) | Qué ve / puede / no puede / firewall por rol. |
| Desiciones de producto | [decisiones-de-producto.csv](decisiones-de-producto.csv) | Decisiones por etapa (nombre de pestaña con typo *Desiciones* en Excel). |
| Flujos por etapa | [flujos-por-etapa.csv](flujos-por-etapa.csv) | Flujo de producto y roles por etapa. |
| UserStories por rol | [user-stories-por-rol.csv](user-stories-por-rol.csv) | Historias C-01… OP-05. |
| UserStories transversales | [user-stories-transversales.csv](user-stories-transversales.csv) | Historias X-01… que aplican a todos los roles. |
| US prioritarias MVP | [us-prioritarias-mvp.csv](us-prioritarias-mvp.csv) | Priorización para MVP. |

Lectura en Markdown auxiliar: [role-experience-matrix.md](role-experience-matrix.md), [matriz-comportamiento-por-rol.md](matriz-comportamiento-por-rol.md), [decisiones-de-producto.md](decisiones-de-producto.md), [flujos-por-etapa.md](flujos-por-etapa.md), [user-stories-transversales.md](user-stories-transversales.md), [us-prioritarias-mvp.md](us-prioritarias-mvp.md).
"""
    dest.write_text(body, encoding="utf-8")


def regenerate_markdown() -> None:
    full = OUT_DIR / "role-model-permission-matrix-full.csv"
    if full.is_file():
        write_roles_por_etapa(full, OUT_DIR / "roles-por-etapa.md")
        write_user_stories_matriz(full, OUT_DIR / "user-stories-matriz.md")
        print("Regenerated roles-por-etapa.md, user-stories-matriz.md")

    us_csv = OUT_DIR / "user-stories-por-rol.csv"
    if us_csv.is_file():
        write_user_stories_por_rol_md(us_csv, OUT_DIR / "user-stories-por-rol.md")
        print("Regenerated user-stories-por-rol.md")

    csv_to_md_table(
        "Role experience Matrix",
        OUT_DIR / "role-experience-matrix.csv",
        OUT_DIR / "role-experience-matrix.md",
        "Derivado de la pestaña homónima del libro Excel.",
    )
    csv_to_md_table(
        "Matriz de comportamiento por rol",
        OUT_DIR / "matriz-comportamiento-por-rol.csv",
        OUT_DIR / "matriz-comportamiento-por-rol.md",
        "Qué ve, qué puede hacer, límites y trazabilidad por rol.",
    )
    csv_to_md_table(
        "Decisiones de producto por etapa",
        OUT_DIR / "decisiones-de-producto.csv",
        OUT_DIR / "decisiones-de-producto.md",
        "Respuesta consolidada, decisión de producto, UX e implicaciones por etapa del ciclo.",
    )
    csv_to_md_table(
        "Flujos por etapa",
        OUT_DIR / "flujos-por-etapa.csv",
        OUT_DIR / "flujos-por-etapa.md",
        "Flujo de producto, objetivo, roles principales y salida por etapa.",
    )
    csv_to_md_table(
        "User stories transversales",
        OUT_DIR / "user-stories-transversales.csv",
        OUT_DIR / "user-stories-transversales.md",
        "Historias que aplican a todos los roles.",
    )
    csv_to_md_table(
        "US prioritarias MVP",
        OUT_DIR / "us-prioritarias-mvp.csv",
        OUT_DIR / "us-prioritarias-mvp.md",
        "Lista priorizada para alcance MVP.",
    )
    write_excel_sheets_index(OUT_DIR / "excel-sheets.md")
    print("Regenerated auxiliary .md and excel-sheets.md")


SHEET_FILES: list[tuple[str, str, bool]] = [
    ("Role Model + Permission Matrix", "role-model-permission-matrix-full.csv", False),
    ("Role experience Matrix", "role-experience-matrix.csv", False),
    ("Matriz de comportamiento por ro", "matriz-comportamiento-por-rol.csv", False),
    ("Desiciones de producto", "decisiones-de-producto.csv", False),
    ("Flujos por etapa", "flujos-por-etapa.csv", False),
    ("UserStories por rol", "user-stories-por-rol.csv", True),
    ("UserStories transversales", "user-stories-transversales.csv", False),
    ("US prioritarias MVP", "us-prioritarias-mvp.csv", False),
    ("V1 Role Model + Permission Matr", "v1-role-model-permission-matrix.csv", False),
]


def main() -> None:
    if not XLSX.is_file():
        print(f"No existe {XLSX}", file=sys.stderr)
        raise SystemExit(1)

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    names = set(wb.sheetnames)
    for sheet, fname, fill_rol in SHEET_FILES:
        if sheet not in names:
            print(f"Aviso: pestaña no encontrada {sheet!r}", file=sys.stderr)
            continue
        ws = wb[sheet]
        m = crop(sheet_matrix(ws))
        if fill_rol:
            m = fill_rol_column_user_stories(m)
        out = OUT_DIR / fname
        write_csv(out, m)
        print(f"Wrote {out.relative_to(REPO)} ({len(m)} rows)")
    wb.close()

    # permission-matrix.csv = misma forma que V1 (sin columnas UX) — copiar desde export V1
    v1 = OUT_DIR / "v1-role-model-permission-matrix.csv"
    perm = OUT_DIR / "permission-matrix.csv"
    if v1.is_file():
        perm.write_text(v1.read_text(encoding="utf-8"), encoding="utf-8")
        print(f"Synced {perm.relative_to(REPO)} from v1 export")

    regenerate_markdown()


if __name__ == "__main__":
    main()
